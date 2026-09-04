"""
Enhanced Excel Document Processor
Erweiterte Excel-Verarbeitung mit Tabellen-Visualisierung und Struktur-Erkennung.
"""

import logging
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime

try:
    import pandas as pd
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logging.warning("pandas/openpyxl nicht verfügbar. Excel-Verarbeitung eingeschränkt.")

logger = logging.getLogger(__name__)


class EnhancedExcelProcessor:
    """Verarbeitet Excel-Dateien mit erweiterten Features."""
    
    def __init__(self):
        self.supported_extensions = ['.xlsx', '.xls', '.xlsm']
        self.max_rows_per_sheet = 10000  # Limit für Performance
        
    def can_process(self, file_path: str) -> bool:
        """Prüft, ob die Datei verarbeitet werden kann."""
        if not EXCEL_AVAILABLE:
            return False
        return Path(file_path).suffix.lower() in self.supported_extensions
    
    def process(self, file_path: str, include_formulas: bool = False) -> Dict[str, Any]:
        """
        Verarbeitet eine Excel-Datei mit erweiterten Features.
        
        Args:
            file_path: Pfad zur Excel-Datei
            include_formulas: Wenn True, werden Formeln extrahiert (langsamer)
            
        Returns:
            Dict mit:
            - text: Formatierter Text (Markdown-Tabellen)
            - metadata: Metadaten
            - sheets: Liste von Sheet-Daten
            - summary: Zusammenfassung der Struktur
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("pandas/openpyxl sind nicht installiert")
            
        try:
            # Excel-Datei laden
            excel_file = pd.ExcelFile(file_path)
            
            # Metadaten
            metadata = self._extract_metadata(file_path, excel_file)
            
            # Sheets verarbeiten
            sheets_data = []
            all_text = []
            
            for sheet_name in excel_file.sheet_names:
                sheet_data = self._process_sheet(
                    excel_file, 
                    sheet_name, 
                    include_formulas=include_formulas
                )
                sheets_data.append(sheet_data)
                
                # Text für RAG
                sheet_text = f"\n\n--- Sheet: {sheet_name} ---\n"
                sheet_text += sheet_data['markdown_table']
                if sheet_data['summary']:
                    sheet_text += f"\n\nSummary: {sheet_data['summary']}"
                all_text.append(sheet_text)
            
            combined_text = "\n".join(all_text)
            
            # Gesamt-Summary
            summary = self._create_summary(sheets_data)
            
            return {
                'text': combined_text,
                'metadata': metadata,
                'sheets': sheets_data,
                'summary': summary,
                'success': True
            }
            
        except Exception as e:
            logger.error(f"Fehler beim Verarbeiten von {file_path}: {e}")
            return {
                'text': '',
                'metadata': {'error': str(e)},
                'sheets': [],
                'summary': '',
                'success': False
            }
    
    def _extract_metadata(self, file_path: str, excel_file: pd.ExcelFile) -> Dict[str, Any]:
        """Extrahiert Metadaten."""
        file_path_obj = Path(file_path)
        
        metadata = {
            'file_name': file_path_obj.name,
            'file_path': str(file_path),
            'file_size': file_path_obj.stat().st_size,
            'sheet_count': len(excel_file.sheet_names),
            'sheet_names': excel_file.sheet_names,
            'processed_at': datetime.now().isoformat(),
            'processor': 'EnhancedExcelProcessor'
        }
        
        # Versuche zusätzliche Metadaten aus openpyxl zu extrahieren
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            props = wb.properties
            metadata.update({
                'title': props.title or 'Unbekannt',
                'author': props.creator or 'Unbekannt',
                'created': props.created.isoformat() if props.created else None,
                'modified': props.modified.isoformat() if props.modified else None
            })
            wb.close()
        except Exception as e:
            logger.warning(f"Konnte erweiterte Metadaten nicht extrahieren: {e}")
        
        return metadata
    
    def _process_sheet(
        self, 
        excel_file: pd.ExcelFile, 
        sheet_name: str, 
        include_formulas: bool = False
    ) -> Dict[str, Any]:
        """Verarbeitet ein einzelnes Sheet."""
        try:
            # DataFrame laden
            df = pd.read_excel(excel_file, sheet_name=sheet_name, nrows=self.max_rows_per_sheet)
            
            # Markdown-Tabelle erstellen
            markdown_table = self._dataframe_to_markdown(df)
            
            # Statistiken
            stats = {
                'rows': len(df),
                'columns': len(df.columns),
                'column_names': list(df.columns),
                'numeric_columns': len(df.select_dtypes(include=['number']).columns),
                'text_columns': len(df.select_dtypes(include=['object']).columns),
                'missing_values': int(df.isna().sum().sum())
            }
            
            # Summary erstellen
            summary = self._create_sheet_summary(df, sheet_name)
            
            # Formeln extrahieren (optional, performance-intensiv)
            formulas = []
            if include_formulas:
                formulas = self._extract_formulas(excel_file.io, sheet_name)
            
            return {
                'sheet_name': sheet_name,
                'markdown_table': markdown_table,
                'stats': stats,
                'summary': summary,
                'formulas': formulas
            }
            
        except Exception as e:
            logger.error(f"Fehler beim Verarbeiten von Sheet {sheet_name}: {e}")
            return {
                'sheet_name': sheet_name,
                'markdown_table': '',
                'stats': {},
                'summary': f'Fehler: {str(e)}',
                'formulas': []
            }
    
    def _dataframe_to_markdown(self, df: pd.DataFrame, max_rows: int = 100) -> str:
        """Konvertiert DataFrame zu Markdown-Tabelle."""
        if len(df) > max_rows:
            df_display = df.head(max_rows)
            footer = f"\n\n... ({len(df) - max_rows} weitere Zeilen nicht angezeigt)"
        else:
            df_display = df
            footer = ""
        
        # Pandas to_markdown (wenn verfügbar)
        try:
            markdown = df_display.to_markdown(index=False)
            return markdown + footer
        except (AttributeError, ImportError):
            # Fallback: Manuell erstellen
            return self._manual_markdown_table(df_display) + footer
    
    def _manual_markdown_table(self, df: pd.DataFrame) -> str:
        """Erstellt Markdown-Tabelle manuell (Fallback)."""
        lines = []
        
        # Header
        header = "| " + " | ".join(str(col) for col in df.columns) + " |"
        separator = "|" + "|".join(["---"] * len(df.columns)) + "|"
        lines.append(header)
        lines.append(separator)
        
        # Rows
        for _, row in df.iterrows():
            row_text = "| " + " | ".join(str(val) for val in row) + " |"
            lines.append(row_text)
        
        return "\n".join(lines)
    
    def _create_sheet_summary(self, df: pd.DataFrame, sheet_name: str) -> str:
        """Erstellt eine Zusammenfassung des Sheets."""
        summary_parts = [
            f"Sheet '{sheet_name}' enthält {len(df)} Zeilen und {len(df.columns)} Spalten."
        ]
        
        # Spalten auflisten
        if len(df.columns) <= 10:
            cols = ", ".join(str(c) for c in df.columns)
            summary_parts.append(f"Spalten: {cols}")
        
        # Numerische Spalten: Stats
        numeric_cols = df.select_dtypes(include=['number']).columns
        if len(numeric_cols) > 0:
            summary_parts.append(f"Numerische Spalten: {len(numeric_cols)}")
        
        return " ".join(summary_parts)
    
    def _extract_formulas(self, file_path: str, sheet_name: str) -> List[Dict[str, Any]]:
        """Extrahiert Formeln aus einem Sheet (optional, langsam)."""
        formulas = []
        try:
            wb = openpyxl.load_workbook(file_path, data_only=False)
            ws = wb[sheet_name]
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formulas.append({
                            'cell': cell.coordinate,
                            'formula': cell.value
                        })
            
            wb.close()
        except Exception as e:
            logger.warning(f"Fehler beim Extrahieren von Formeln: {e}")
        
        return formulas
    
    def _create_summary(self, sheets_data: List[Dict[str, Any]]) -> str:
        """Erstellt Gesamt-Zusammenfassung."""
        total_rows = sum(sheet['stats'].get('rows', 0) for sheet in sheets_data)
        total_cols = sum(sheet['stats'].get('columns', 0) for sheet in sheets_data)
        
        summary = f"Excel-Datei mit {len(sheets_data)} Sheet(s), "
        summary += f"{total_rows} Zeilen total, {total_cols} Spalten total."
        
        return summary


# Singleton-Instanz
_excel_processor = None

def get_enhanced_excel_processor() -> EnhancedExcelProcessor:
    """Liefert Singleton-Instanz des Enhanced Excel-Processors."""
    global _excel_processor
    if _excel_processor is None:
        _excel_processor = EnhancedExcelProcessor()
    return _excel_processor
